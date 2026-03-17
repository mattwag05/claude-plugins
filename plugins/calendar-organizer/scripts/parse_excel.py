#!/usr/bin/env python3
"""Parse Excel/CSV files and output structured cell data as JSON.

Usage:
    python parse_excel.py /path/to/file.xlsx [--sheet "Sheet Name"]
    python parse_excel.py /path/to/file.csv

Outputs JSON with structure:
{
    "filename": "file.xlsx",
    "sheets": ["Sheet1", "Sheet2"],
    "active_sheet": "Sheet1",
    "rows": [
        {"row": 1, "cells": [{"col": "A", "value": "Date", "col_idx": 1}, ...]},
        ...
    ],
    "dimensions": {"min_row": 1, "max_row": 50, "min_col": 1, "max_col": 10}
}
"""

import json
import sys
import csv
from pathlib import Path


def parse_excel(filepath: str, sheet_name: str | None = None) -> dict:
    """Parse an Excel file into structured JSON."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        cells = []
        for cell in row:
            if cell.value is not None:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                cells.append({
                    "col": col_letter,
                    "col_idx": cell.column,
                    "value": str(cell.value).strip(),
                })
        if cells:
            rows.append({"row": row[0].row, "cells": cells})

    return {
        "filename": Path(filepath).name,
        "sheets": wb.sheetnames,
        "active_sheet": ws.title,
        "rows": rows,
        "dimensions": {
            "min_row": ws.min_row,
            "max_row": ws.max_row,
            "min_col": ws.min_column,
            "max_col": ws.max_column,
        },
    }


def parse_csv(filepath: str) -> dict:
    """Parse a CSV file into structured JSON."""
    rows = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            cells = []
            for col_idx, value in enumerate(row, start=1):
                if value.strip():
                    col_letter = chr(64 + col_idx) if col_idx <= 26 else f"C{col_idx}"
                    cells.append({
                        "col": col_letter,
                        "col_idx": col_idx,
                        "value": value.strip(),
                    })
            if cells:
                rows.append({"row": row_idx, "cells": cells})

    max_col = max((c["col_idx"] for r in rows for c in r["cells"]), default=1)
    return {
        "filename": Path(filepath).name,
        "sheets": ["Sheet1"],
        "active_sheet": "Sheet1",
        "rows": rows,
        "dimensions": {
            "min_row": 1,
            "max_row": len(rows),
            "min_col": 1,
            "max_col": max_col,
        },
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: parse_excel.py <file> [--sheet <name>]", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]
    sheet_name = None
    if "--sheet" in sys.argv:
        idx = sys.argv.index("--sheet")
        if idx + 1 < len(sys.argv):
            sheet_name = sys.argv[idx + 1]

    path = Path(filepath)
    if not path.exists():
        print(f"Error: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        result = parse_excel(filepath, sheet_name)
    elif ext == ".csv":
        result = parse_csv(filepath)
    else:
        print(f"Error: Unsupported file type: {ext}", file=sys.stderr)
        sys.exit(1)

    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
